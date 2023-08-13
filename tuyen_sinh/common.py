import openpyxl

class ClassKhoi:
    def __init__(self, name):
        self.name = name
        self.list_subject = []
        self.he_so = {}
        self.gap_point = 0
    
    def print_info(self):
        print("the Subject Combination is", self.name)
        print("list subject is", self.list_subject)

class ClassMajor:
    def __init__(self, major_id):
        self.major_id = major_id
        self.maToHop = []
        self.primary_khoi = ClassKhoi("None")
        self.is_add = True
    
    def add_khoi(self, khoi):
        self.is_add == True
        if 0 == len(self.maToHop):
            self.maToHop.append(khoi)
        else:
            for temp in self.maToHop:
                if temp.name == khoi.name:
                    self.is_add = False
                    print(khoi, "is exist in", self.major_id)
                    break
            if self.is_add == True:
                self.maToHop.append(khoi)
                
    def update_primary_khoi(self, primary_khoi):
            self.primary_khoi=primary_khoi

    def print_info(self):
        print("the Major ID is", self.major_id)
        print("primary_khoi is", self.primary_khoi.name)
        print("maToHop is")
        for temp in self.maToHop:
            print("\t",temp.name,"list_subject",temp.list_subject,"he_so",temp.he_so,"gap_point",temp.gap_point)
        print()

class Student:
    def __init__(self, id_student):
        self.id_student = id_student
        self.list_NV = []
        self.major = ClassMajor()

    def print_info(self):
        print("the id_student is", self.id_student)

class ExcelHandler:
    class SheetIndexMapping:
        def __init__(self):
            self.index_colum = 1
            self.index_row = 1
            self.max_index_colum = 1
            self.max_index_row = 1
        
        def inc_index_column(self):
            self.index_colum += 1
            if(self.max_index_colum < self.index_colum):
                self.max_index_colum = self.index_colum

        def inc_index_row(self):
            self.index_row += 1
            if(self.max_index_row < self.index_row):
                self.max_index_row = self.index_row

    def __init__(self, name_file):
        self.name_file = name_file
        print("Open excell file name", name_file)
        self.workbook = openpyxl.load_workbook(self.name_file)

        self.current_sheet = self.workbook.active
        self.listSheet = {}

        for temp in self.workbook.sheetnames:
            # print("Exist sheet",temp)
            self.listSheet[temp] = ExcelHandler.SheetIndexMapping()

    def chosse_current_sheet(self, sheet_to_focus):
        if sheet_to_focus not in self.workbook.sheetnames:
            print("sheet name",sheet_to_focus,"is not exist")
            return
        
        for temp in self.workbook.sheetnames:
            if temp == sheet_to_focus:
                self.workbook.active.title = temp
                self.current_sheet = self.workbook.active
                return self.workbook.active.title

    def print_info(self):
        print("file name is {self.name_file}")
        print("current sheet name is {self.current_sheet}")
    
    def print_data_collum(self, my_collum_value):
        for col in self.current_sheet.iter_rows(1, self.current_sheet.max_row):
            print(col[my_collum_value].value)

    def print_data_row(self, my_row_value):
        for col in self.current_sheet.iter_cols(1, self.current_sheet.max_column):
            print(my_row_value[col].value)
    
    def add_sheet(self, sheet_name_to_add):
        self.workbook.create_sheet(sheet_name_to_add)
        self.listSheet[sheet_name_to_add] = ExcelHandler.SheetIndexMapping()
        print("created sheet", sheet_name_to_add)

    def remove_sheet(self, sheet_name_to_remove):
        if sheet_name_to_remove in self.workbook.sheetnames:
            sheet_to_remove = self.workbook[sheet_name_to_remove]
            self.workbook.remove(sheet_to_remove)
            print(f"Sheet '{sheet_name_to_remove}' has been removed from the Excel file.")
        else:
            print(f"Sheet '{sheet_name_to_remove}' does not exist in the Excel file.")
    
    def find_colum_index_with_content(self, content):
        temp = 0
        for col in self.current_sheet.iter_cols(1, self.current_sheet.max_column):
            if (col[0].value == content):
                print("find column number contain", content,"in header of sheet",self.current_sheet.title, "value =",temp)
                return temp
            temp += 1

    def find_row_index_with_content(self, content):
        print("find row number contain", content,"in sheet",self.current_sheet.title)
        temp = 0
        for row in self.current_sheet.iter_rows(1, self.current_sheet.max_row):
            if (row[0].value == content):
                return temp
            temp += 1
    
    def save_file(self):
        self.workbook.save(self.name_file)

    def sort_inc_base_on_column_index(self, column_index):
        # Get all rows (excluding the first row, which is typically a header)
        data_rows = list(self.current_sheet.iter_rows(min_row=2, values_only=True))

        # Create a list of rows where each row is a tuple containing values from all columns
        rows_with_values = []
        for row in data_rows:
            # Pad the row with None values to account for missing columns
            padded_row = row + (None,) * (self.current_sheet.max_column - len(row))
            rows_with_values.append(padded_row)

        # Sort the rows based on the values in the specified column in descending order
        sorted_rows = sorted(rows_with_values, key=lambda row: row[column_index - 1], reverse=True)

        # Clear the original sheet content
        for row in self.current_sheet.iter_rows(min_row=2, max_row=self.current_sheet.max_row, min_col=1, max_col=self.current_sheet.max_column):
            for cell in row:
                cell.value = None

        # Write the sorted rows back to the sheet
        for row_index, data_row in enumerate(sorted_rows, start=2):
            for col_index, value in enumerate(data_row, start=1):
                self.current_sheet.cell(row=row_index, column=col_index, value=value)
    
    def remove_row_index(self, row_index_to_remove):
        # Delete the specified row
        self.current_sheet.delete_rows(row_index_to_remove)

    def remove_row_index_with_sheet(self, sheet_to_remove_row, row_index_to_remove):
        # Delete the specified row
        temp = self.current_sheet
        self.chosse_current_sheet(sheet_to_remove_row)
        self.current_sheet.delete_rows(row_index_to_remove)
        self.chosse_current_sheet(temp)
        