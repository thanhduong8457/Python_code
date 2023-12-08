from datetime import datetime
import numpy as np
import openpyxl     # Import for working with excel

def check_excel(workbook_name, sheet_name):
    wb = openpyxl.load_workbook(workbook_name) # define variable to load the dataframe
    
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == sheet_name:
            break
            
    wb.active = s
    sub_wb = wb.active # define variable to read sheet
    
    # wb.create_sheet('output')
    wb.save(sheet_name)

# month/day/year
# 3/7/2023

year = '2023'
month = '03'
day = '07'

period = 7

def is_a_busday()
    start = year + '-' + month + '-' + day

    temp = start.split('-')

    print(start)
    print(temp)

    if not np.is_busday(start):
        print("Not a Business day")
        return True
    else:
        print("Business day")
        return False
